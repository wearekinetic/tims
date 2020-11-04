from openpyxl.styles import *
from openpyxl.utils import get_column_letter


def set_style(cells, *styles):
    for row in cells:
        for cell in row:
            for style in styles:
                if isinstance(style, Font):
                    cell.font = style
                elif isinstance(style, PatternFill):
                    cell.fill = style


def xlrange(row0, row1, col0, col1):
    return f'{col0}{row0}:{col1}{row1}'


def style_current_row(sheet, *styles):
    row_idx = sheet.max_row 
    range_ = xlrange(row_idx, row_idx, 'A', 'I')
    row = sheet[range_]
    set_style(row, *styles)



bold_font = Font(name='Calibri', size=11, bold=True)
bold_fill = PatternFill('solid', fgColor='d0d0d0')

split_fill = PatternFill('solid', fgColor='ddddff')
repositioning_fill = PatternFill('solid', fgColor='ffdddd')