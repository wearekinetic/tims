import openpyxl
from openpyxl.utils import get_column_letter

from timesheet import SplitShiftItem, MinimumShiftAdjustment, Alert, Adjustment
from styles import *
from driver import Driver
import settings


# constants
SUMMARY_HEADER = ['ID', 'Driver', 'Hours', 'Paid Break', 'Unpaid Break', 'Total Payable']
SUMMARY_COLUMN_WIDTHS = [15, 40, 15, 15, 15, 15, 15]

SHEET_HEADER = ['Type', 'Start', 'Pickup', 'Destination', 'End', 'Hours', 'Paid Break', 'Unpaid Break', 'Total Payable']
SHEET_COLUMN_WIDTHS = [15, 15, 40, 40, 15, 15, 15, 15, 15]


def export(path, timesheets):
    workbook = openpyxl.Workbook()

    summary = workbook.active  
    summary.title = 'Summary'

    summary.append(SUMMARY_HEADER)
    row = summary['A1:F1']
    set_style(row, bold_font, bold_fill)

    # set column widths
    for idx, width in enumerate(SUMMARY_COLUMN_WIDTHS):
        summary.column_dimensions[get_column_letter(idx + 1)].width = width

    # detail sheets
    for timesheet in timesheets:
        driver = timesheet.driver
        print (driver.name)
        sheet = workbook.create_sheet(title=driver.code)

        # set column widths
        for idx, width in enumerate(SHEET_COLUMN_WIDTHS):
            sheet.column_dimensions[get_column_letter(idx + 1)].width = width
        
        # output header
        sheet.append(SHEET_HEADER)
        style_current_row(sheet, bold_font, bold_fill)

        # output shifts
        shifts = timesheet.shifts[:]
        while shifts:
            shift = shifts.pop(0)

            for item in shift.items:
                if isinstance(item, Adjustment):
                    sheet.append([])
                
                sheet.append(item.output())

                # bold text for adjustments
                if isinstance(item, Adjustment):
                    style_current_row(sheet, bold_font)

                if hasattr(item, 'colour'):
                    fill = PatternFill('solid', fgColor=item.colour)
                    style_current_row(sheet, fill)

                    if isinstance(item, Alert):
                         style_current_row(sheet, bold_font)

            if shifts:
                sheet.append([])
                sheet.append(timesheet.split.output())

                fill = PatternFill('solid', fgColor=settings.SPLIT_SHIFT_COLOUR)
                style_current_row(sheet, fill)
                sheet.append([])

        # summary
        summary_row = [driver.code, driver.name, timesheet.total_worked_hours.to_time(), timesheet.total_paid_breaks.to_time(), timesheet.total_unpaid_breaks.to_time(), timesheet.total_payable_hours.to_time()]
        summary.append(summary_row)
  
        # output footer/totals
        sheet.append([])
        footer = ['Total', None, None, None, timesheet.total_time.to_time(), timesheet.total_worked_hours.to_time(), timesheet.total_paid_breaks.to_time(), timesheet.total_unpaid_breaks.to_time(), timesheet.total_payable_hours.to_time()]
        sheet.append(footer)

        # style total row
        style_current_row(sheet, bold_font, bold_fill)
        
    workbook.save(path)


     